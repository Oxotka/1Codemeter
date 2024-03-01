import settings
import re
import os
import git
import copy
import openpyxl
from openpyxl.styles import Font
from tqdm import tqdm


def color_text(text, color):
    return '<span style="color:{color}">{text}</span>'.format(color=color, text=text)


def write_line(file, content, tag=''):
    if tag == '':
        file.write(content)
    else:
        if '#' in tag:
            # Headings should be surrounded by blank lines
            file.write('\n')
        file.write('{tag} {content}'.format(tag=tag, content=content))
        if '#' in tag:
            file.write('\n')
    file.write('\n')


def open_details(title, file):
    write_line(file, '<details>', '')
    write_line(file, '  <summary><i>{title}</i></summary>'.format(title=title), '')
    write_line(file, '', '')


def close_details(file):
    write_line(file, '</details>', '')
    write_line(file, '', '')


def write_title(title, file):
    write_line(file, title, '')
    write_line(file, '', '')


def icon_md(name):
    return '<img title="{name}" align=center width=16 height=16 src="icons/{name}.png"> '.format(name=name)


def print_authors(authors, lines_info, file):

    if len(lines_info) == 0:
        return

    new_lines_info = {}
    for author in lines_info:
        name = '{name} <{email}>'.format(name=authors.get(author), email=author)
        new_lines_info[name] = lines_info.get(author)

    sorted_lines_info = dict(sorted(new_lines_info.items()))

    number = 1
    for author in sorted_lines_info:
        author_info = sorted_lines_info.get(author)
        write_line(file, '{num}. {name} {insert} {delete}'.format(
            num=number,
            name=author,
            insert=color_text('+{}'.format(author_info.get('insert', 0)), 'rgb(0,128,0)'),
            delete=color_text('-{}'.format(author_info.get('delete', 0)), 'rgb(255,0,0)')))
        number += 1
    write_line(file, '')


def print_subsystem(subsystems, file):
    if len(subsystems) > 0:
        subsystems = sorted(subsystems)
        write_line(file, "##### Подсистемы")
        write_line(file, '')
        for subsystem in subsystems:
            write_line(file, subsystem, '-')
        write_line(file, '')


class StructureOfCodemeter:
    def __init__(self):
        self.path_to_repo = os.path.normpath(settings.path_to_repo())  # Путь до локального репозитория
        self.name_of_src = os.path.normpath(settings.name_of_src())  # Имя в папке с конфигурацией
        self.date_since = settings.date_since()  # Дата, с которой начинаем читать коммиты
        self.date_before = settings.date_before()  # Дата, до которой читаем коммиты
        self.exclude_subsystems = settings.exclude_subsystems()  # Исключаемые подсистемы
        self.include_subsystems = settings.include_subsystems()  # Включаемые подсистемы
        self.configuration_name = ''  # Имя конфигурации из файлов конфигурации
        self.commits = []  # Все подходящие коммиты, между датами date_since и date_before
        self.subsystems = []  # Служебный массив всех подсистем. Собирается из файлов конфигурации
        self.subsystem_by_object = {}  # Служебный словарь подсистем. {type: {object: [subsystem1]}}
        self.authors = {}  # Авторы в формате {емайл : имя}. Заполняется автоматически при чтении коммитов
        self.structure_of_conf = {}  # Итоговая структура конфигурации

    def get_structure_subsystem(self):
        if len(self.subsystems) == 0:
            return
        for subsystem in self.subsystems:
            self.get_subsystem_info(subsystem)

    def get_subsystem_info(self, subsystem):
        for info in subsystem:
            if len(subsystem.get(info).get('subsystems')) > 0:
                for inner_subsystem in subsystem.get(info).get('subsystems'):
                    self.get_subsystem_info(inner_subsystem)

            for content in subsystem.get(info).get('contents'):
                elements = single_to_plural(content).split('.')
                if len(elements) != 2:
                    continue
                type_name = elements[0]
                object_name = elements[1]
                type_info = self.subsystem_by_object.get(type_name, {})
                object_info = type_info.get(object_name, [])
                if info not in object_info:
                    object_info.append(info)
                type_info[object_name] = object_info
                self.subsystem_by_object[type_name] = type_info

    def get_commits_info(self):
        repo = git.Repo(self.path_to_repo)
        commits = list(repo.iter_commits("master"))
        print('')
        print('Statistics collection has started')
        print('Please wait. It may take a long time...')
        print('')
        print('The number of all commits in the repository: {len}'.format(len=len(commits)))
        print('')
        if self.date_before is not None and self.date_since is not None:
            print('Processing is performed only between these dates: {since} and {before}'.format(
                since=self.date_since.date(), before=self.date_before.date()))
            print('Other commits will be skipped and the process may stop before the progress bar completes.')

        elif self.date_since is not None:
            print('Processing is performed only since {since}'.format(
                since=self.date_since.date()))
            print('Other commits will be skipped and the process may stop before the progress bar completes.')
        elif self.date_before is not None:
            print('Processing is performed only before {before}'.format(
                before=self.date_since.date()))
            print('Other commits will be skipped and the process may stop before the progress bar completes.')

        with tqdm(total=len(commits), desc='Get commits', ncols=100, colour='green') as pbar:
            for commit in commits:
                pbar.update(1)

                if self.date_before is not None \
                        and commit.committed_datetime.timestamp() > self.date_before.timestamp():
                    continue

                if self.date_since is not None \
                        and self.date_since.timestamp() >= commit.committed_datetime.timestamp():
                    print('Date of commit ({commit}) are earlier then date_since ({since})'.format(
                        commit=commit.committed_datetime.date(), since=self.date_since.date()))
                    print('It is okay, we stop get commit and go forward')
                    break

                for file in commit.stats.files:
                    # find only in chosen configuration in settings and .bsl files
                    if os.path.normpath(file).startswith(os.path.join(self.name_of_src, 'src')) \
                            and file.endswith('bsl'):
                        stat = {'date': commit.committed_datetime.date(),
                                'file': file,
                                'insert': commit.stats.files.get(file).get('insertions'),
                                'delete': commit.stats.files.get(file).get('deletions'),
                                'email': commit.author.email}
                        self.commits.append(stat)
                        self.authors[commit.author.email] = commit.author.name

    def summarize_info_to_contents(self):
        summarized = {}
        if len(self.commits) == 0:
            return summarized
        for commit in self.commits:
            file = commit.get('file')
            email = commit.get('email')
            file_info = summarized.get(file)
            if file_info is None:
                file_info = {email: {'insert': 0, 'delete': 0}}
            email_info = file_info.get(email)
            if email_info is None:
                email_info = {'insert': 0, 'delete': 0}
            email_info['insert'] = email_info.get('insert') + commit.get('insert')
            email_info['delete'] = email_info.get('delete') + commit.get('delete')
            file_info[email] = email_info
            summarized[file] = file_info

        return summarized

    def structure_by_content_and_subsystem(self):
        summarized = self.summarize_info_to_contents()
        structure_of_configuration = {}
        with tqdm(total=len(summarized), desc='Summarize info', ncols=100, colour='green') as pbar:
            for file in summarized:
                pbar.update(1)
                email_info = summarized.get(file)
                # example: DemoConfDT/src/AccumulationRegisters/Взаиморасчеты/Forms/ТекущиеВзаиморасчеты/Module.bsl
                file = os.path.normpath(file)
                file = file.replace(os.path.join(self.name_of_src, 'src'), '')
                parts_of_name = file.split(os.path.sep)
                type_name = parts_of_name[1]  # example: AccumulationRegisters
                object_name = parts_of_name[2]  # example: Взаиморасчеты
                type_info = copy.deepcopy(structure_of_configuration.get(type_name, {}))
                object_info = copy.deepcopy(type_info.get(object_name, {}))
                info = object_info
                for i in range(3, len(parts_of_name)):
                    inner_info = info.get(parts_of_name[i])
                    if inner_info is None:
                        info.update({parts_of_name[i]: {}})
                        inner_info = info.get(parts_of_name[i])
                    info = inner_info
                    if i == len(parts_of_name) - 1:
                        info.update(email_info)
                type_info[object_name] = object_info
                skip = False
                subsystem_type = self.subsystem_by_object.get(type_name, {})
                subsystems = subsystem_type.get(object_name, [])
                if len(self.include_subsystems) > 0 and len(self.subsystem_by_object) > 0:
                    it_is_include = False
                    for include in self.include_subsystems:
                        if include != "" and include in subsystems:
                            it_is_include = True
                            break
                    skip = not it_is_include

                if len(self.exclude_subsystems) > 0 and len(self.subsystem_by_object) > 0:
                    for exclude in self.exclude_subsystems:
                        if exclude != "":
                            for object_subsystem in subsystems:
                                if exclude in object_subsystem:
                                    skip = True
                                    break
                if skip:
                    continue

                # stats by obj
                for email_info_by_author in email_info:
                    authors = copy.deepcopy(object_info.get('authors', {}))
                    if authors.get(email_info_by_author) is None:
                        upd_author = email_info.get(email_info_by_author, {})
                    else:
                        upd_author = copy.deepcopy(authors.get(email_info_by_author))
                        upd_author['insert'] = upd_author.get('insert', 0) + email_info.get(
                            email_info_by_author).get('insert', 0)
                        upd_author['delete'] = upd_author.get('delete', 0) + email_info.get(
                            email_info_by_author).get('delete', 0)
                    authors[email_info_by_author] = upd_author
                    object_info['authors'] = authors

                    authors = copy.deepcopy(type_info.get('authors', {}))
                    if authors.get(email_info_by_author) is None:
                        upd_author = email_info.get(email_info_by_author, {})
                    else:
                        upd_author = copy.deepcopy(authors.get(email_info_by_author))
                        upd_author['insert'] = upd_author.get('insert', 0) + email_info.get(
                            email_info_by_author).get('insert', 0)
                        upd_author['delete'] = upd_author.get('delete', 0) + email_info.get(
                            email_info_by_author).get('delete', 0)
                    authors[email_info_by_author] = upd_author
                    type_info['authors'] = authors

                    # common info about stats
                    structure_authors = structure_of_configuration.get('authors', {})
                    if structure_authors.get(email_info_by_author) is None:
                        structure_author = email_info.get(email_info_by_author)
                    else:
                        structure_author = copy.deepcopy(structure_authors.get(email_info_by_author))

                        structure_author['insert'] = structure_author.get('insert', 0) + email_info.get(
                            email_info_by_author).get('insert', 0)
                        structure_author['delete'] = structure_author.get('delete', 0) + email_info.get(
                            email_info_by_author).get('delete', 0)
                    structure_authors[email_info_by_author] = structure_author

                    structure_of_configuration['authors'] = structure_authors
                type_info = dict(sorted(type_info.items()))
                structure_of_configuration.update({type_name: type_info})
        self.structure_of_conf = structure_of_configuration

    def save_to_markdown(self):
        if len(self.structure_of_conf) == 0:
            return

        with tqdm(total=len(self.structure_of_conf), desc='Save to markdown', ncols=100, colour='green') as pbar:
            with open('stats_info.md', 'w', encoding='utf-8') as result_file:
                write_line(result_file, self.configuration_name, '#')
                open_details('Отборы:', result_file)
                if self.date_since is not None \
                        and self.date_before is not None:
                    write_line(result_file, 'Коммиты собраны начиная с {since} по {before}'.format(
                        since=self.date_since.date(), before=self.date_before.date()))
                    write_line(result_file, '')
                elif self.date_since is not None:
                    write_line(result_file,
                               'Коммиты собраны начиная с {since}'.format(since=self.date_since.date()))
                    write_line(result_file, '')
                elif self.date_before is not None:
                    write_line(result_file, 'Коммиты собраны по {before}'.format(before=self.date_before.date()))
                    write_line(result_file, '')
                if len(self.include_subsystems) > 0:
                    write_line(result_file, 'Отбор по этим подсистемам:')
                    for subsystem in self.include_subsystems:
                        write_line(result_file, subsystem, '-')
                    write_line(result_file, '')
                if len(self.exclude_subsystems) > 0:
                    write_line(result_file, 'Исключая эти подсистемы:')
                    for subsystem in self.exclude_subsystems:
                        write_line(result_file, subsystem, '-')
                    write_line(result_file, '')
                close_details(result_file)

                write_line(result_file, 'Разработчики:', '##')
                print_authors(self.authors, self.structure_of_conf.get('authors'), result_file)

                write_line(result_file, 'Объекты:', '##')
                for object_name in self.structure_of_conf:
                    pbar.update(1)
                    if object_name == 'authors' or object_name == 'Configuration':
                        continue
                    else:
                        subsystem_obj = self.subsystem_by_object.get(object_name, {})
                        write_line(result_file, icon_md(object_name) + object_name, '###')
                        open_details('Подробнее', result_file)
                        types = self.structure_of_conf.get(object_name)
                        for type_name in types:
                            if type_name == 'authors':
                                continue
                            else:
                                write_line(result_file, icon_md(object_name) + type_name, '####')
                                object_info = types.get(type_name)
                                print_authors(self.authors, object_info.get('authors'), result_file)
                                subsystem_type = subsystem_obj.get(type_name, [])
                                print_subsystem(subsystem_type, result_file)
                                if object_name == 'Catalogs' or object_name == 'DataProcessors' \
                                        or object_name == 'Documents' or object_name == 'Reports':
                                    open_details('Еще', result_file)
                                    if object_info.get('ObjectModule.bsl') is not None:
                                        write_title('##### Модуль объекта', result_file)
                                        lines_info = object_info.get('ObjectModule.bsl')
                                        print_authors(self.authors, lines_info, result_file)

                                    if object_info.get('ManagerModule.bsl') is not None:
                                        write_title('##### Модуль менеджера', result_file)
                                        lines_info = object_info.get('ManagerModule.bsl')
                                        print_authors(self.authors, lines_info, result_file)

                                    if object_info.get('Forms') is not None:
                                        write_title('##### Формы', result_file)
                                        forms_info = object_info.get('Forms')
                                        for form in forms_info:
                                            write_title(form, result_file)
                                            lines_info = forms_info.get(form).get('Module.bsl')
                                            print_authors(self.authors, lines_info, result_file)
                                    close_details(result_file)

                                elif object_name == 'InformationRegisters' or object_name == 'AccumulationRegisters':
                                    open_details('Еще', result_file)
                                    if object_info.get('RecordSetModule.bsl') is not None:
                                        write_title('##### Модуль записи', result_file)
                                        lines_info = object_info.get('RecordSetModule.bsl')
                                        print_authors(self.authors, lines_info, result_file)
                                    if object_info.get('Forms') is not None:
                                        write_title('##### Формы', result_file)
                                        forms_info = object_info.get('Forms')
                                        for form in forms_info:
                                            write_title(form, result_file)
                                            lines_info = forms_info.get(form).get('Module.bsl')
                                            print_authors(self.authors, lines_info, result_file)
                                    close_details(result_file)

                        close_details(result_file)

    def save_to_excel(self):
        if len(self.structure_of_conf) == 0:
            return
        wb = openpyxl.Workbook()
        wb.create_sheet(title='Все данные', index=0)
        sheet = wb['Все данные']
        title_font = Font(name='Arial', size=18)
        bold_font = Font(bold=True)
        row_title = 2
        sheet['B{}'.format(row_title)] = self.configuration_name
        sheet['B{}'.format(row_title)].font = title_font
        row_title += 1
        if self.date_since is not None \
                and self.date_before is not None:
            sheet['B{}'.format(row_title)] = 'Коммиты с {since} по {before}'.format(
                since=self.date_since.date(), before=self.date_before.date())
            row_title += 1
        elif self.date_since is not None:
            sheet['B{}'.format(row_title)] = 'Коммиты с {since}'.format(since=self.date_since.date())
            row_title += 1
        elif self.date_before is not None:
            sheet['B{}'.format(row_title)] = 'Коммиты по {before}'.format(before=self.date_before.date())
            row_title += 1
        if len(self.include_subsystems) > 0:
            sheet['B{}'.format(row_title)] = 'Отбор по этим подсистемам: {subsystems}'.format(
                subsystems=', '.join(self.include_subsystems))
            row_title += 1
        if len(self.exclude_subsystems) > 0:
            sheet['B{}'.format(row_title)] = 'Исключая эти подсистемы: {subsystems}'.format(
                subsystems=', '.join(self.exclude_subsystems))
            row_title += 1
        row_title += 1
        column_titles = {'type': 2, 'object': 3, 'subsystem': 4, 'author': 5, 'email': 6, 'insert': 7, 'delete': 8}
        for col in column_titles:
            cell = sheet.cell(row=row_title, column=column_titles[col])
            cell.value = col
            cell.font = bold_font
        row = row_title
        with tqdm(total=len(self.structure_of_conf), desc='Save to Excel', ncols=100, colour='green') as pbar:
            for type_name in self.structure_of_conf:
                pbar.update(1)
                if type_name == 'authors' or type_name == 'Configuration':
                    continue
                else:
                    subsystem_obj = self.subsystem_by_object.get(type_name, {})
                    objects = self.structure_of_conf.get(type_name)
                    for object_name in objects:
                        if object_name == 'authors':
                            continue
                        else:
                            object_info = objects.get(object_name)
                            authors_info = object_info.get('authors')
                            for author in authors_info:
                                row += 1
                                sheet.cell(row=row, column=column_titles['type']).value = type_name
                                sheet.cell(row=row, column=column_titles['object']).value = object_name
                                sheet.cell(row=row, column=column_titles['subsystem']).value = \
                                    ', '.join(subsystem_obj.get(object_name, []))
                                sheet.cell(row=row, column=column_titles['email']).value = author
                                sheet.cell(row=row, column=column_titles['author']).value = self.authors[author]
                                sheet.cell(row=row, column=column_titles['insert']).value = \
                                    authors_info.get(author).get('insert')
                                sheet.cell(row=row, column=column_titles['delete']).value = \
                                    authors_info.get(author).get('delete')

        wb.save('stats.xlsx')


def single_to_plural(content):
    if content.startswith('FilterCriterion'):
        return content.replace("FilterCriterion", "FilterCriteria")
    if content.startswith('ChartOfCharacteristicTypes'):
        return content.replace("ChartOfCharacteristicTypes", "ChartsOfCharacteristicTypes")
    else:
        return content.replace(".", "s.")


def path_to_object(content):
    content = single_to_plural(content)
    parts_of_name = content.split('.')
    return content.replace(".", "/") + "/" + parts_of_name[1] + '.mdo'


def get_structure_of_configuration():
    structure_of_codemeter = StructureOfCodemeter()
    if structure_of_codemeter.path_to_repo == '':
        print('Path to repo is empty. Please check settings.py')
        return

    path = os.path.join(structure_of_codemeter.path_to_repo, structure_of_codemeter.name_of_src)
    configuration = os.path.normpath('src/Configuration/Configuration.mdo')
    subsystem_path = os.path.normpath('src/Subsystems/')

    reg_exp_pattern_subsystem = '(?<=<subsystems>Subsystem.).*?(?=</subsystems>)'
    reg_exp_pattern_content = '(?<=<content>).*?(?=</content>)'
    reg_exp_pattern_name = '(?<=<value>).*?(?=</value>)'

    path_to_configuration = os.path.join(path, configuration)
    if not os.path.isfile(path_to_configuration):
        print('Configuration file is not found by path - {path}. Please check settings.py'.format(
            path=path_to_configuration))
        return

    # get upper subsystems
    upper_subsystems = []
    with open(path_to_configuration, mode='r', encoding='utf8') as f:
        file = f.read().encode('utf-8').decode('utf-8')
        m = re.search(reg_exp_pattern_name, file)
        if not (m is None):
            structure_of_codemeter.configuration_name = m.group()

        for subsystem in re.findall(reg_exp_pattern_subsystem, file):
            upper_subsystems.append(subsystem)

    # get info about subsystems
    for subsystem in upper_subsystems:
        path_to_dir_subsystem = os.path.join(path, subsystem_path, subsystem)
        info_subsystem = info_about_subsystems(subsystem, '', path_to_dir_subsystem, reg_exp_pattern_content)
        structure_of_codemeter.subsystems.append(info_subsystem)

    structure_of_codemeter.get_structure_subsystem()
    structure_of_codemeter.get_commits_info()
    structure_of_codemeter.structure_by_content_and_subsystem()

    return structure_of_codemeter


def info_about_subsystems(subsystem, upper_subsystem, path, reg_exp_pattern_content):
    subsystem_name = os.path.join(path, subsystem + '.mdo')
    full_subsystem = subsystem
    if upper_subsystem != '':
        full_subsystem = f'{upper_subsystem}.{subsystem}'

    contents = []
    subsystems = []
    if os.path.isdir(os.path.join(path, 'Subsystems')):
        inner_subsystems = [f for f in os.listdir(os.path.join(path, 'Subsystems')) if f != '.DS_Store']
        for inner_subsystem in inner_subsystems:
            inner_subsystem_path = os.path.join(path, 'Subsystems', inner_subsystem)
            inner_info = info_about_subsystems(inner_subsystem,
                                               full_subsystem, inner_subsystem_path, reg_exp_pattern_content)
            subsystems.append(inner_info)

    with open(subsystem_name, mode='r', encoding='utf8') as f:
        file = f.read().encode('utf-8').decode('utf-8')
        for content in re.findall(reg_exp_pattern_content, file):
            contents.append(content)

    return {full_subsystem: {'subsystems': subsystems, 'contents': contents}}


def get_statistics():
    structure = get_structure_of_configuration()
    if structure is None:
        return

    save_to_md = settings.save_to_md()
    save_to_excel = settings.save_to_excel()
    if save_to_md:
        structure.save_to_markdown()
    if save_to_excel:
        structure.save_to_excel()
    print('')

    if len(structure.structure_of_conf) == 0:
        print('Nothing was found. Check the settings')
        print('For example: settings.date_since() and settings.date_before()')
    else:
        print('Statistics are collected!')
        if save_to_md and save_to_excel:
            print('Please check result files: stats_info.md and stats.xlsx')
        elif save_to_md:
            print('Please check result file: stats_info.md')
        elif save_to_excel:
            print('Please check result file: stats.xlsx')
        else:
            print('Result file has not been saved. Please check settings.py - save_to_md() and save_to_excel()')


if __name__ == '__main__':
    get_statistics()

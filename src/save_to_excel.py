import openpyxl
from openpyxl.styles import Font
from tqdm import tqdm


def save(configuration, path='result/stats.xlsx'):
    if len(configuration.structure_of_conf) == 0:
        return
    wb = openpyxl.Workbook()
    wb.create_sheet(title='Все данные', index=0)
    sheet = wb['Все данные']
    title_font = Font(name='Arial', size=18)
    bold_font = Font(bold=True)
    row_title = 2
    sheet['B{}'.format(row_title)] = configuration.configuration_name
    sheet['B{}'.format(row_title)].font = title_font
    row_title += 1
    if configuration.date_since is not None \
            and configuration.date_before is not None:
        sheet['B{}'.format(row_title)] = 'Коммиты с {since} по {before}'.format(
            since=configuration.date_since.date(), before=configuration.date_before.date())
        row_title += 1
    elif configuration.date_since is not None:
        sheet['B{}'.format(row_title)] = 'Коммиты с {since}'.format(since=configuration.date_since.date())
        row_title += 1
    elif configuration.date_before is not None:
        sheet['B{}'.format(row_title)] = 'Коммиты по {before}'.format(before=configuration.date_before.date())
        row_title += 1
    if len(configuration.include_subsystems) > 0:
        sheet['B{}'.format(row_title)] = 'Отбор по этим подсистемам: {subsystems}'.format(
            subsystems=', '.join(configuration.include_subsystems))
        row_title += 1
    if len(configuration.exclude_subsystems) > 0:
        sheet['B{}'.format(row_title)] = 'Исключая эти подсистемы: {subsystems}'.format(
            subsystems=', '.join(configuration.exclude_subsystems))
        row_title += 1
    row_title += 1
    column_titles = {'type': 2, 'object': 3, 'subsystem': 4, 'author': 5, 'email': 6, 'insert': 7, 'delete': 8}
    for col in column_titles:
        cell = sheet.cell(row=row_title, column=column_titles[col])
        cell.value = col
        cell.font = bold_font
    row = row_title
    with tqdm(total=len(configuration.structure_of_conf), desc='Save to Excel', ncols=100, colour='green') as pbar:
        for type_name in configuration.structure_of_conf:
            pbar.update(1)
            if type_name == 'authors' or type_name == 'Configuration':
                continue
            else:
                subsystem_obj = configuration.subsystem_by_object.get(type_name, {})
                objects = configuration.structure_of_conf.get(type_name)
                for object_name in objects:
                    if object_name == 'authors':
                        continue
                    else:
                        object_info = objects.get(object_name)
                        authors_info = object_info.get('authors')
                        for author in authors_info:
                            row += 1
                            sheet.cell(row=row, column=column_titles['type']).value = type_name
                            sheet.cell(row=row, column=column_titles['object']).value = object_name
                            sheet.cell(row=row, column=column_titles['subsystem']).value = \
                                ', '.join(subsystem_obj.get(object_name, []))
                            sheet.cell(row=row, column=column_titles['email']).value = author
                            sheet.cell(row=row, column=column_titles['author']).value = configuration.authors[author]
                            sheet.cell(row=row, column=column_titles['insert']).value = \
                                authors_info.get(author).get('insert')
                            sheet.cell(row=row, column=column_titles['delete']).value = \
                                authors_info.get(author).get('delete')

    wb.save(path)
